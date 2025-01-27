import sys
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QVBoxLayout,
    QPushButton,
    QFileDialog,
    QTableWidget,
    QTableWidgetItem,
    QWidget,
)
import openpyxl
from Categories import Categories
from Data_Plots import plot_spending

class MA_Finance(QMainWindow):
    def __init__(self):
        super().__init__()

        # Set up the main window
        self.setWindowTitle("MA_Finance")
        self.setGeometry(200, 200, 800, 600)

        # Main layout
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)

        # Add a button to load Excel files
        self.load_button = QPushButton("Import Statement")
        self.load_button.clicked.connect(self.load_excel_file)
        self.layout.addWidget(self.load_button)

        self.plot_button = QPushButton('Open Plot Window', self)
        self.plot_button.clicked.connect(self.open_plot_window)
        self.layout.addWidget(self.plot_button)

        # Add a table to display Excel content
        self.table = QTableWidget()
        self.layout.addWidget(self.table)

    def load_excel_file(self):
        """
         # Open a file dialog to select an Excel file
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Import Statement", "", "Excel Files (*.xlsx *.xls *.csv)"
        )
        if file_path:
            #self.load_format_content(file_path)
        """
        self.load_format_content(r"C:\Users\jasle\OneDrive\Documents\GitHub\Budget_Plannig\2024_Chequing_Transactions_only.xlsx")

    def load_format_content(self, file_path):
        # Load the Excel file
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            # Get the number of rows and columns
            rows = sheet.max_row
            cols = sheet.max_column

            # Set the table dimensions
            self.table.setRowCount(rows)
            self.table.setColumnCount(cols)

            # Populate the table with Excel data
            for row in range(1, rows + 1):
                for col in range(1, cols + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    self.table.setItem(row - 1, col - 1, QTableWidgetItem(str(cell_value)))
            
            #make first row items headers
            for col in range(1, cols + 1):
                header = sheet.cell(row=1, column=col).value
                self.table.setHorizontalHeaderItem(col-1, QTableWidgetItem(str(header)))
            
            #Remove the first row
            self.table.removeRow(0)

            # Add a column to the table named category and move it to the first column
            self.table.insertColumn(0)
            self.table.setHorizontalHeaderItem(0, QTableWidgetItem("Category"))

            # Remove columns that are not needed
            columns_to_keep = ["category", "date", "sub-description", "amount", "balance"]
            columns_to_remove = []
            for col in range(self.table.columnCount()):
                header_item = self.table.horizontalHeaderItem(col)
                #if header_item and header_item.text().lower() in ["status", "filter", "description", "type of transaction", "None"]:
                if header_item and header_item.text().lower() not in columns_to_keep:
                    columns_to_remove.append(col)

            # Remove the columns in reverse order to avoid shifting issues
            for col in sorted(columns_to_remove, reverse=True):
                self.table.removeColumn(col)
            
            # Check if any cells in sub-description column are empty and set them to "Not Specified"
            for row in range(self.table.rowCount()):
                item = self.table.item(row, 2)
                if item is None or item == "None" or not item.text().strip():
                    self.table.setItem(row, 2, QTableWidgetItem("Not Specified"))
                    self.table.setItem(row, 0, QTableWidgetItem("Not Specified"))

            Categories(self.table)
            
            # Remove columns that are not needed
            self.table.removeColumn(5)
            self.table.removeColumn(5)

        except Exception as e:
            print(f"Error loading Excel file: {e}")

    def open_plot_window(self):
        self.plot_window = plot_spending(self.table, self)
        self.plot_window.show()
        
# Run the application
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MA_Finance()
    window.show()
    sys.exit(app.exec_())
